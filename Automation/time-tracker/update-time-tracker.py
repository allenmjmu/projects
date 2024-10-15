import openpyxl
from datetime import datetime, timedelta
import os
from copy import copy

def copy_cell_format(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = source_cell.number_format  # Direct assignment
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

def auto_size_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
      
# Load the Excel workbook
file_path = '/Users/allen/Desktop/time-tracker.xlsx'
if not os.access(file_path, os.W_OK):
    raise PermissionError(f"No write permission for the file: {file_path}")

workbook = openpyxl.load_workbook(file_path)

# Print all worksheet names to verify the correct name
print("Available worksheets:", workbook.sheetnames)

# Determine the original worksheet name (most recent one)
def extract_start_date(name):
    return datetime.strptime(name.split('-')[0], '%m%d%y')

original_worksheet_name = max(workbook.sheetnames, key=extract_start_date)

# Calculate the new worksheet name based on the current date
current_date = extract_start_date(original_worksheet_name)
new_start_date = current_date + timedelta(days=7)
new_end_date = new_start_date + timedelta(days=6)
new_worksheet_name = f"{new_start_date.strftime('%m%d%y')}-{new_end_date.strftime('%m%d%y')}"

# Create a new worksheet
new_worksheet = workbook.create_sheet(new_worksheet_name)
print(f"New worksheet '{new_worksheet_name}' created.")

# Select the original worksheet
if original_worksheet_name in workbook.sheetnames:
    original_worksheet = workbook[original_worksheet_name]
else:
    raise ValueError(f"Worksheet '{original_worksheet_name}' does not exist in the workbook.")

# Set the number of days to offset the dates
date_offset = 7

# Copy the headers from the original worksheet to the new worksheet
for j in range(1, original_worksheet.max_column + 1):
    source_cell = original_worksheet.cell(row=1, column=j)
    target_cell = new_worksheet.cell(row=1, column=j)
    target_cell.value = source_cell.value
    copy_cell_format(source_cell, target_cell)

# Get the maximum row number
last_row = original_worksheet.max_row

# Loop through each row and copy/update the dates
for i in range(2, last_row + 1):  # Assuming row 1 is the header
    # Copy the previous week's data
    for j in range(1, original_worksheet.max_column + 1):
        source_cell = original_worksheet.cell(row=i, column=j)
        target_cell = new_worksheet.cell(row=i, column=j)
        target_cell.value = source_cell.value
        copy_cell_format(source_cell, target_cell)

    # Update the dates in columns A, H, J by adding the offset
    for col in ['A', 'H', 'J']:
        cell_value = original_worksheet[f'{col}{i}'].value
        if isinstance(cell_value, datetime):
            new_date = cell_value + timedelta(days=date_offset)
            new_worksheet[f'{col}{i}'].value = new_date
            new_worksheet[f'{col}{i}'].number_format = 'MM/DD/YY'

# Auto-size the columns in the new worksheet
auto_size_columns(new_worksheet)

# Save the modified workbook
try:
    workbook.save(file_path)
    print(f"Workbook saved with new worksheet '{new_worksheet_name}'.")
except PermissionError as e:
    print(f"Failed to save the workbook: {e}")
